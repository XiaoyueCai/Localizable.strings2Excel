#!/usr/bin/python
# -*- coding: UTF-8 -*-

import openpyxl


class XlsFileUtil:
    """xls file util"""

    def __init__(self, filePath):
        self.filePath = filePath
        self.data = openpyxl.load_workbook(filePath)

    def getAllTables(self):
        names = self.data.sheetnames
        return [self.data[name] for name in names]

    def getTableByIndex(self, index):
        if 0 <= index < len(self.data.sheetnames):
            return self.data[self.data.sheetnames[index]]
        else:
            print("XlsFileUtil error -- getTable:index")

    def getTableByName(self, name):
        return self.data[name]